from crewai_sellers_flow.ports.whatsapp_lka_repository import WhatsappLKARepository
from pydantic import BaseModel
import os
import requests
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.client_credential import ClientCredential
import pandas as pd
from crewai_sellers_flow.adapters.lka_repository import LKA_INSTANCES
from collections import defaultdict
class WhatsappLKASharepoint(WhatsappLKARepository): 
    """Adapter para o repositório de LKA dos sellers."""

    def __init__(self):
        self.site_url = os.getenv("SHAREPOINT_ZLABS_SITE_URL")
        self.client_id = os.getenv("SHAREPOINT_CLIENT_ID")
        self.client_secret = os.getenv("SHAREPOINT_CLIENT_SEC")
        self.file_id = os.getenv("SHAREPOINT_ZLABS_FILE_ID")
        self.ctx = None
        self._sellers_index = None

    class LKA(BaseModel):
        instance: str
        token: str
        phone_seller: str

    def setup(self):
        if self._sellers_index is None:
            file_name = self._download_file("Mailing Sellers.xlsx")
            self._build_sellers_index(file_name)

    def _connect_sharepoint(self):
        """
        Estabelece conexão com o SharePoint.
        
        Returns:
            bool: True se a conexão foi bem sucedida, False caso contrário
        """
        credentials = ClientCredential(self.client_id, self.client_secret)
        self.ctx = ClientContext(self.site_url).with_credentials(credentials)    

    def _download_file(self, file_name: str) -> str:
        """
        Faz download de um arquivo do SharePoint.
        
        Args:
            file_name (str): Nome do arquivo para salvar localmente
            
        Returns:
            str: Caminho do arquivo baixado
        """
        try:
            if not self.ctx:
                self._connect_sharepoint()

            temp_file = f"/tmp/{file_name}"

            file = self.ctx.web.get_file_by_id(self.file_id)

            os.makedirs(os.path.dirname(temp_file), exist_ok=True)

            with open(temp_file, "wb") as local_file:
                file.download(local_file).execute_query()

            print(f"Arquivo baixado com sucesso: {temp_file}")
            return temp_file
        except Exception as e:
            print(f"Erro ao fazer download do arquivo: {str(e)}")
            raise Exception(f"Erro ao fazer download do arquivo: {str(e)}")        

    def _read_excel(self, path: str) -> pd.DataFrame | None:
        try:
            # Usar pandas com engine='openpyxl' mas sem data_only
            df = pd.read_excel(path, sheet_name=0, engine='openpyxl')
            
            # Tratar NaN
            df = df.fillna('')
            
            print(f"DataFrame carregado com {len(df)} linhas e {len(df.columns)} colunas")
            print(f"Colunas: {list(df.columns)}")
            
            return df
        except Exception as e:
            print(f"Erro ao ler arquivo Excel: {e}")
            raise Exception(f"Erro ao ler arquivo Excel: {e}")
    
    def _build_sellers_index(self, path: str):
        import openpyxl
        from math import isnan

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        ws = wb["Aux"]
        lka = defaultdict(str)
        for row in ws.iter_rows(min_row=2, values_only=True):
            lka[row[4]] = row[5]

        ws = wb["BASE OFICIALLL"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        want = ["POC_ID", "Celular", "Responsável", "Sellers_weekly"]
        idx = {name: header.index(name) for name in want if name in header}
        index = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            poc = row[idx["POC_ID"]]
            if poc is None:
                continue
            try:
                poc_int = int(poc)
            except Exception:
                continue
            responsavel = row[idx["Responsável"]]
            celular = row[idx["Celular"]]
            seller_weekly = row[idx["Sellers_weekly"]]
            index[poc_int] = {
                "responsavel": "" if responsavel is None else str(responsavel),
                "celular": "" if celular is None else str(celular),
                "lka_celular": lka.get(responsavel,""),
                "seller_weekly": True if seller_weekly is None else bool(seller_weekly)
            }
        self._sellers_index = index

    def get_lka(self, seller_id: int) -> LKA:
        self.setup()

        seller = self._sellers_index.get(int(seller_id))
        if not seller:
            raise ValueError(f"POC com ID {seller_id} não encontrado no arquivo Mailing Sellers.xlsx")

        instance = next(
            (item for item in LKA_INSTANCES if item["phoneNumber"] == str(seller["lka_celular"])),
            None,
        )
        if not instance:
            raise ValueError(f"LKA não encontrado com fone {seller['lka_celular']}")
        
        if not seller["seller_weekly"]: 
            raise ValueError(f"POC com ID {seller_id} está desabilitado para Sellers Weekly")

        return self.LKA(
            instance=instance["Instance"],
            token=instance["Token"],
            phone_seller=str(seller["celular"]),
        )


if __name__ == "__main__":
    whatsapp_lka_sharepoint = WhatsappLKASharepoint()
    print(whatsapp_lka_sharepoint.get_lka(122852))
    print(whatsapp_lka_sharepoint.get_lka(31370))
    print(whatsapp_lka_sharepoint.get_lka(76404))
    print(whatsapp_lka_sharepoint.get_lka(1))
